#  Hãy tạo 2 files excel và đặt vào ổ đĩa nào đó, sử dụng đường dẫn tuyệt đối để load 2 file này lên python.
# Giả sử files có tên là file1.xlsx và file2.xlsx
# Trong file1.xlsx sẽ có dữ liệu sheet1 gồm 2 cột A và B bất kỳ dữ liệu gì
# file2.xlsx có sheet1 nhưng hoàn toàn không có dữ liệu nào.
# Yêu cầu:
# Hãy copy nội dung cột B của file 1 vào cột A của file 2 trong sheet1 nhưng chỉ copy 5 dòng đầu tiên.
from openpyxl import load_workbook

# Mở file excel
workbook_file1 = load_workbook(filename="D:\\Personal\\IMIC\\file1.xlsx")
workbook_file2 = load_workbook(filename="D:\\Personal\\IMIC\\file2.xlsx")

sheet1_file1 = workbook_file1["Sheet1"]
sheet1_file2 = workbook_file2["Sheet1"]

for r in range(1, 10):
    for c in range(1, 3):
        sheet1_file2.cell(row=r, column=c).value = sheet1_file1.cell(row=r, column=c).value

# Lưu file excel
workbook_file2.save(filename="D:\\Personal\\IMIC\\file2.xlsx")

workbook_file1.close()
workbook_file2.close()
